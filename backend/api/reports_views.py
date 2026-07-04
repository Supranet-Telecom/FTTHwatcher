"""
FTTH Watcher — Geração de relatórios (PDF e Excel) do panorama de uma cidade.

O frontend envia os dados já calculados (KPIs, ranking, segmento) e as imagens
dos gráficos (PNG em base64). Este módulo só monta o documento — assim o relatório
fica idêntico ao que aparece na tela, sem reimplementar a agregação em Python.
"""

import base64
import io
from datetime import datetime

from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView


# Ordem e rótulos dos gráficos no relatório
CHART_ORDER = [
    ("acessos", "Acessos por Empresa"),
    ("participacao", "Participação de Mercado (%)"),
    ("crescimento", "Crescimento Mensal por Empresa"),
    ("segmento", "Perfil de Clientes — PF vs PJ"),
]


def _decode_png(data_url: str) -> bytes | None:
    """Converte 'data:image/png;base64,XXXX' em bytes. Retorna None se inválido."""
    if not data_url or "," not in data_url:
        return None
    try:
        return base64.b64decode(data_url.split(",", 1)[1])
    except Exception:
        return None


def _fmt_int(v):
    try:
        return f"{int(round(float(v))):,}".replace(",", ".")
    except (TypeError, ValueError):
        return "—"


def _fmt_pct(v):
    if v is None:
        return "—"
    try:
        return f"{float(v):+.1f}%"
    except (TypeError, ValueError):
        return "—"


def _titulo(payload) -> str:
    cidade = payload.get("cidade") or payload.get("uf") or "Brasil"
    return f"Panorama de Mercado — {cidade}"


# ---------------------------------------------------------------------------
# PDF (reportlab)
# ---------------------------------------------------------------------------
def build_pdf(payload) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title=_titulo(payload),
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#0f172a"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#334155"))
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748b"))

    story = []
    story.append(Paragraph(_titulo(payload), h1))
    periodo = payload.get("periodo", "—")
    gerado = datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph(f"Período de referência: {periodo} &nbsp;•&nbsp; Gerado em {gerado}", small))
    story.append(Spacer(1, 0.5 * cm))

    # --- KPIs ---
    k = payload.get("kpis", {})
    kpi_rows = [
        ["Total de acessos", _fmt_int(k.get("total_mercado")),
         "Acessos Supranet", _fmt_int(k.get("supra_acessos"))],
        ["Share Supranet", _fmt_pct(k.get("supra_share")).replace("+", ""),
         "Posição Supranet", f"{k.get('supra_rank')}º" if k.get("supra_rank") else "—"],
        ["Concorrentes ativos", str(k.get("num_concorrentes", "—")),
         "Cresc. Supranet (mês)", _fmt_pct(k.get("supra_cresc"))],
        ["Cresc. médio concorrentes", _fmt_pct(k.get("cresc_medio_conc")), "", ""],
    ]
    kpi_table = Table(kpi_rows, colWidths=[4.5 * cm, 4 * cm, 4.5 * cm, 4 * cm])
    kpi_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#64748b")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#64748b")),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(Paragraph("Indicadores", h2))
    story.append(kpi_table)
    story.append(Spacer(1, 0.4 * cm))

    # --- Ranking ---
    ranking = payload.get("ranking", [])
    if ranking:
        story.append(Paragraph("Ranking de Empresas", h2))
        data = [["#", "Empresa", "Acessos", "Share", "Cresc. mês"]]
        for i, r in enumerate(ranking, 1):
            data.append([
                str(i), r.get("nome", "—"),
                _fmt_int(r.get("acessos")),
                _fmt_pct(r.get("share")).replace("+", ""),
                _fmt_pct(r.get("crescimento")),
            ])
        tbl = Table(data, colWidths=[1 * cm, 7 * cm, 3 * cm, 2.5 * cm, 3 * cm])
        style = [
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        # Destaca a linha da Supranet em laranja
        for i, r in enumerate(ranking, 1):
            if (r.get("nome") or "").lower() == "supranet":
                style.append(("TEXTCOLOR", (0, i), (-1, i), colors.HexColor("#ea580c")))
                style.append(("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"))
        tbl.setStyle(TableStyle(style))
        story.append(tbl)
        story.append(Spacer(1, 0.4 * cm))

    # --- Gráficos ---
    charts = payload.get("charts", {})
    content_width = A4[0] - 3 * cm
    for key, label in CHART_ORDER:
        png = _decode_png(charts.get(key, ""))
        if not png:
            continue
        try:
            img = Image(io.BytesIO(png))
            ratio = img.imageHeight / img.imageWidth
            img.drawWidth = content_width
            img.drawHeight = content_width * ratio
            # KeepTogether mantém título + gráfico na mesma página (evita título órfão)
            story.append(KeepTogether([
                Paragraph(label, h2),
                Spacer(1, 0.15 * cm),
                img,
                Spacer(1, 0.4 * cm),
            ]))
        except Exception:
            continue

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel (openpyxl)
# ---------------------------------------------------------------------------
def build_xlsx(payload) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    orange = Font(color="EA580C", bold=True)

    # --- Resumo ---
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = _titulo(payload)
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Período: {payload.get('periodo', '—')} — Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=9, color="64748B")

    k = payload.get("kpis", {})
    kpis = [
        ("Total de acessos", _fmt_int(k.get("total_mercado"))),
        ("Acessos Supranet", _fmt_int(k.get("supra_acessos"))),
        ("Share Supranet", _fmt_pct(k.get("supra_share")).replace("+", "")),
        ("Posição Supranet", f"{k.get('supra_rank')}º" if k.get("supra_rank") else "—"),
        ("Concorrentes ativos", k.get("num_concorrentes", "—")),
        ("Cresc. Supranet (mês)", _fmt_pct(k.get("supra_cresc"))),
        ("Cresc. médio concorrentes", _fmt_pct(k.get("cresc_medio_conc"))),
    ]
    row = 4
    for label, val in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(color="64748B")
        ws.cell(row=row, column=2, value=val).font = Font(bold=True)
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    # --- Ranking ---
    ranking = payload.get("ranking", [])
    if ranking:
        wr = wb.create_sheet("Ranking")
        headers = ["#", "Empresa", "Acessos", "Share (%)", "Cresc. mês (%)"]
        for c, h in enumerate(headers, 1):
            cell = wr.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for i, r in enumerate(ranking, 1):
            wr.cell(row=i + 1, column=1, value=i)
            nome_cell = wr.cell(row=i + 1, column=2, value=r.get("nome"))
            wr.cell(row=i + 1, column=3, value=r.get("acessos"))
            wr.cell(row=i + 1, column=4, value=r.get("share"))
            wr.cell(row=i + 1, column=5, value=r.get("crescimento"))
            if (r.get("nome") or "").lower() == "supranet":
                nome_cell.font = orange
        for col, width in zip("ABCDE", [5, 32, 14, 12, 14]):
            wr.column_dimensions[col].width = width

    # --- Segmento PF/PJ ---
    segmento = payload.get("segmento", [])
    if segmento:
        wsg = wb.create_sheet("Segmento PF-PJ")
        for c, h in enumerate(["Empresa", "Pessoa Física", "Pessoa Jurídica", "Total"], 1):
            cell = wsg.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        for i, s in enumerate(segmento, 1):
            pf = s.get("pf", 0) or 0
            pj = s.get("pj", 0) or 0
            wsg.cell(row=i + 1, column=1, value=s.get("nome"))
            wsg.cell(row=i + 1, column=2, value=pf)
            wsg.cell(row=i + 1, column=3, value=pj)
            wsg.cell(row=i + 1, column=4, value=pf + pj)
        for col, width in zip("ABCD", [32, 16, 16, 12]):
            wsg.column_dimensions[col].width = width

    # --- Gráficos ---
    charts = payload.get("charts", {})
    imgs = [(label, _decode_png(charts.get(key, ""))) for key, label in CHART_ORDER]
    imgs = [(label, png) for label, png in imgs if png]
    if imgs:
        wg = wb.create_sheet("Gráficos")
        anchor_row = 1
        for label, png in imgs:
            wg.cell(row=anchor_row, column=1, value=label).font = Font(bold=True)
            try:
                xi = XLImage(io.BytesIO(png))
                # Reduz para caber na planilha (largura ~ 700px)
                scale = min(1.0, 700 / xi.width)
                xi.width = int(xi.width * scale)
                xi.height = int(xi.height * scale)
                wg.add_image(xi, f"A{anchor_row + 1}")
                anchor_row += int(xi.height / 18) + 3
            except Exception:
                anchor_row += 2

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# View
# ---------------------------------------------------------------------------
class CityReportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # NÃO usar "format" — é palavra reservada do DRF (negociação de conteúdo).
        fmt = request.query_params.get("tipo", "pdf").lower()
        payload = request.data or {}
        cidade = (payload.get("cidade") or payload.get("uf") or "brasil").lower().replace(" ", "_")
        periodo = (payload.get("periodo") or "").replace("-", "")

        if fmt == "xlsx":
            content = build_xlsx(payload)
            ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext = "xlsx"
        else:
            content = build_pdf(payload)
            ctype = "application/pdf"
            ext = "pdf"

        resp = HttpResponse(content, content_type=ctype)
        resp["Content-Disposition"] = f'attachment; filename="panorama_{cidade}_{periodo}.{ext}"'
        return resp
